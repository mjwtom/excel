#!/usr/bin/env python

import sys
import os
import datetime
import calendar
from openpyxl import Workbook
from openpyxl import load_workbook


day2list = ['一', '二', '三', '四', '五', '六', '日']
max_row = 20
max_col = 9
row_header_len = 2


def convert(path):
    # 读取excel
    wb = load_workbook(filename=path)
    sheet_ranges = wb['主网']
    # 根据B2处时间获得月份
    month_detect_date_time = sheet_ranges['D2'].value
    print(month_detect_date_time)
    # 获得这个月的总天数
    print('获得月份信息....')
    year = month_detect_date_time.year
    month = month_detect_date_time.month
    print(month)
    num_days = calendar.monthrange(year, month)[1]
    # 生成这个月的所有天
    days = [datetime.date(year, month, day) for day in range(1, num_days + 1)]

    new_wb = Workbook()
    ws1 = new_wb.active
    ws1.title = '主网'
    ws1['A1'] = '月历'
    ws1['A2'] = '周历'

    # 生成月历和周历表头
    print('生成表头......')
    for i in range(0, len(days)):
        col = 2 + i
        ws1.cell(column=col, row=1, value=1+i)
    for i in range(0, len(days)):
        col = 2 + i
        day_index = days[i].weekday()
        ws1.cell(column=col, row=2, value=day2list[day_index])

    base_row_index = 3
    max_row_index = 3
    bounds = []
    for row in sheet_ranges.iter_rows(min_row=2, max_col=max_col, max_row=max_row):
        device = row[2].value
        if not device:
            print('扫描结束....')
            break
        print('发现设备：'+device)
        station = row[1].value
        if station:
            info = station + ':' + device
        else:
            info = device
        info += '。'
        owner = row[7].value
        if owner:
            info += owner
        start_datetime = row[3].value
        end_datetime = row[6].value
        # 开始查找合适插入位置
        search_index = 0
        found = False
        while search_index < len(bounds):
            start, end = bounds[search_index]
            if start_datetime > end or end_datetime < start:
                found = True
                break
            search_index += 1
        if not found:
            bounds.append((start_datetime, end_datetime))
            ws1.cell(column=1, row=search_index + base_row_index, value=search_index+1)
        start_cell_colum = start_datetime.day + 1
        end_cell_colum = end_datetime.day + 1
        ws1.cell(column=start_cell_colum, row=search_index + base_row_index, value=info)
        ws1.merge_cells(start_row=search_index + base_row_index,
                        start_column=start_cell_colum, end_row=search_index + base_row_index,
                        end_column=end_cell_colum)

    print('保存文件......')
    file_name = 'converted-' + os.path.basename(path)
    dir = os.path.dirname(path)
    new_path = os.path.join(dir, file_name)
    new_wb.save(filename=new_path)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('give the file path')
    convert(sys.argv[1])